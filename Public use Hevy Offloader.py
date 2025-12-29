import os
import pandas as pd
import requests
from datetime import datetime
from openpyxl import load_workbook

# --------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------
# Recommended: set HEVY_API_KEY in your environment and use that.
# But keeping your inline key behavior as-is for now.
API_KEY = os.getenv("HEVY_API_KEY") or "INSERT YOUR API KEY HERE"

BASE_URL = "https://api.hevyapp.com"
ROUTINES_ENDPOINT = "/v1/routines"
EXCEL_FILE = "Min-Max_Program_4x.xlsx"
SHEET_NAME = "4x Per Week"

# Set to True to only print what would be created (no API calls)
DRY_RUN = False

HEADERS = {
    "api-key": API_KEY,
    "Content-Type": "application/json",
    "accept": "application/json",
}

# Jeff exercise name -> Hevy exercise_template_id
EXERCISE_MAP = {
    "1-Arm Reverse Pec Deck": "7a0128ce-90ec-45b1-a445-7de6ac03bed0",
    "Alternating DB Curl": "37FCC2BB",
    "Barbell Incline Press": "50DFDFAB",
    "Barbell RDL": "2B4B7310",
    "Bayesian Cable Curl": "234897AB",
    "Cable Crunch": "23A48484",
    "Cable Triceps Kickback": "EC3B69A3",
    "Chest-Supported T-Bar Row": "914F3A96",
    "Close-Grip Lat Pulldown": "4E5257DE",
    "DB Wrist Curl": "1006DF48",
    "DB Wrist Extension": "9202CC23",
    "Dead Hang (optional)": "B9380898",
    "High-Cable Lateral Raise": "DE68C825",
    "Incline DB Y-Raise": "F21D5693",
    "Leg Extension": "75A4F6C4",
    "Leg Press": "C7973E0E",
    "Lying Leg Curl": "B8127AD1",
    "Machine Chest Press": "7EB3F7C3",
    "Machine Hip Thrust": "68CE0B9B",
    "Machine Lateral Raise": "D5D0354D",
    "Machine Shrug": "19A38071",
    "Modified Zottman Curl": "123EE239",
    "Overhead Cable Triceps Extension": "B5EFBF9C",
    "Pull-Up (Wide Grip)": "7C50F118",
    "Squat (Your Choice)": "D04AC939",
    "Standing Calf Raise": "06745E58",
}

REST_LABELS = {"1-2 Rest Days", "Rest Day", "Rest Days"}


# --------------------------------------------------------------------
# HELPERS: Excel parsing
# --------------------------------------------------------------------
def find_all_weeks(df: pd.DataFrame):
    """
    Find all 'Week X' header rows and return list of:
    [(week_number, start_idx, end_idx), ...]
    """
    week_col = "The Min-Max Program"
    week_entries = []

    for idx, val in df[week_col].items():
        if isinstance(val, str) and val.startswith("Week "):
            parts = val.split()
            if len(parts) >= 2 and parts[1].isdigit():
                week_num = int(parts[1])
                week_entries.append((week_num, idx))

    week_ranges = []
    for i, (week_num, start_idx) in enumerate(week_entries):
        end_idx = week_entries[i + 1][1] if i < len(week_entries) - 1 else len(df)
        week_ranges.append((week_num, start_idx, end_idx))

    return week_ranges


def parse_week_days(df_week: pd.DataFrame):
    """
    Given the slice of the sheet for a week (excluding the Week header row),
    return a list of (day_name, [row_indices]) for that week.
    """
    routines = []
    current_day = None
    current_ex_rows = []

    for idx, row in df_week.iterrows():
        day = row["The Min-Max Program"]
        ex = row["Unnamed: 2"]

        # New day label (Full Body, Upper, Lower, Arms/Delts, etc.)
        if isinstance(day, str) and day not in REST_LABELS and not day.startswith("Week"):
            if current_day and current_ex_rows:
                routines.append((current_day, current_ex_rows))
            current_day = day
            current_ex_rows = []

        # Exercise row: has an exercise name
        if isinstance(ex, str) and ex.strip():
            current_ex_rows.append(idx)

    if current_day and current_ex_rows:
        routines.append((current_day, current_ex_rows))

    return routines


def rep_lower_from_value(val):
    """Extract lower rep from rep range (handles Excel date weirdness)."""
    from pandas import Timestamp

    if isinstance(val, (Timestamp, datetime)):
        # rep range like 6-8 got turned into date -> month is lower bound
        return val.month
    if pd.isna(val):
        return None

    s = str(val).strip()
    for sep in ["-", "–", "to"]:
        if sep in s:
            part = s.split(sep)[0].strip()
            if part.isdigit():
                return int(part)

    try:
        return int(float(s))
    except Exception:
        return None


def warmup_count_from_value(val) -> int:
    """
    Convert warm-up sets cell into an integer count.
    Rule: use the LOWER end of a range.
      - "1-2" -> 1
      - "2-3" -> 2
      - 3 -> 3
      - blank/N/A -> 0
    Handles Excel date weirdness similarly (month is lower bound).
    """
    from pandas import Timestamp

    if val is None or pd.isna(val):
        return 0

    if isinstance(val, (Timestamp, datetime)):
        # warm-up like 1-2 got turned into date -> month is lower bound
        return int(val.month)

    s = str(val).strip()
    if not s or s.lower() in {"n/a", "na", "none"}:
        return 0

    for sep in ["-", "–", "to"]:
        if sep in s:
            left = s.split(sep)[0].strip()
            return int(left) if left.isdigit() else 0

    try:
        n = int(float(s))
        return max(n, 0)
    except Exception:
        return 0


def safe_str(val, default="N/A"):
    if val is None or pd.isna(val):
        return default
    s = str(val).strip()
    return s if s else default


def get_exercise_link(ws, df_index: int):
    """
    Using openpyxl, get the hyperlink attached to the exercise cell
    corresponding to this DataFrame row index.
    Assumes header row is row 1 in Excel, df index 0 => row 2.
    """
    excel_row = df_index + 2
    cell = ws.cell(row=excel_row, column=3)  # Column C = exercise name
    return cell.hyperlink.target if cell.hyperlink else None


def extract_exercise_info(df: pd.DataFrame, ws, row_idx: int):
    """Extract all fields we need from one exercise row."""
    row = df.loc[row_idx]

    name = row["Unnamed: 2"]

    warmup_raw = row.get("Unnamed: 4", None)  # warm-up sets cell (may be range/date/blank)
    warmup_count = warmup_count_from_value(warmup_raw)

    working_sets = int(row["Unnamed: 5"]) if not pd.isna(row["Unnamed: 5"]) else 0
    rep_lower = rep_lower_from_value(row["Unnamed: 6"])

    # RIR / Failure columns (as you’ve been using)
    rir1 = safe_str(row.get("Unnamed: 11", None))
    rir2 = safe_str(row.get("Unnamed: 12", None))

    rest = safe_str(row.get("Unnamed: 13", None))
    sub1 = safe_str(row.get("Unnamed: 14", None), default="-")
    sub2 = safe_str(row.get("Unnamed: 15", None), default="-")
    link = get_exercise_link(ws, row_idx)

    # Notes (Option B) — WITHOUT warm-up line now
    # Order you requested:
    # 1) Failure / RIR set 1 and 2
    # 2) Rest
    # 3) Substitution 1 & 2
    # 4) Link attached to exercise name
    notes_lines = [
        f"Set 1 RIR: {rir1}",
        f"Set 2 RIR: {rir2}",
        f"Rest: {rest}",
        f"Sub 1: {sub1}",
        f"Sub 2: {sub2}",
    ]
    if link:
        notes_lines.append(f"Video: {link}")

    notes = "\n".join(notes_lines)

    return {
        "df_index": row_idx,
        "name": name,
        "warmup_count": warmup_count,
        "working_sets": working_sets,
        "rep_lower": rep_lower,
        "rir1": rir1,
        "rir2": rir2,
        "rest": rest,
        "sub1": sub1,
        "sub2": sub2,
        "link": link,
        "notes": notes,
    }


# --------------------------------------------------------------------
# BUILD ROUTINES FOR ALL WEEKS
# --------------------------------------------------------------------
def build_all_routines():
    df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    week_ranges = find_all_weeks(df)

    all_routines = []
    for week_num, start_idx, end_idx in week_ranges:
        week_df = df.iloc[start_idx + 1 : end_idx]  # exclude the "Week X" header row
        day_groups = parse_week_days(week_df)

        for day_name, row_indices in day_groups:
            exercises = [extract_exercise_info(df, ws, idx) for idx in row_indices]
            all_routines.append(
                {
                    "week": week_num,
                    "day_name": day_name,
                    "routine_title": f"Week {week_num} - {day_name}",
                    "exercises": exercises,
                }
            )

    return all_routines


# --------------------------------------------------------------------
# CONVERT TO HEVY PAYLOAD
# --------------------------------------------------------------------
def routine_to_hevy_payload(routine):
    exercises_payload = []

    for ex in routine["exercises"]:
        name = ex["name"]
        hevy_id = EXERCISE_MAP.get(name)
        if not hevy_id:
            print(f"WARNING: No Hevy ID mapped for exercise: {name}")
            continue

        if ex["working_sets"] <= 0:
            continue

        reps = ex["rep_lower"]
        if reps is None:
            print(f"WARNING: Could not parse rep range lower bound for: {name}")
            continue

        sets = []

        # Add warm-up sets FIRST using Hevy set type
        # (Hevy supports set "type"; we use "warmup" here)
        for _ in range(ex["warmup_count"]):
            sets.append(
                {
                    "type": "warmup",
                    "weight_kg": None,
                    "reps": reps,
                    "distance_meters": None,
                    "duration_seconds": None,
                    "custom_metric": None,
                    "rep_range": {"start": reps, "end": None},
                }
            )

        # Add working sets
        for _ in range(ex["working_sets"]):
            sets.append(
                {
                    "type": "normal",
                    "weight_kg": None,
                    "reps": reps,
                    "distance_meters": None,
                    "duration_seconds": None,
                    "custom_metric": None,
                    "rep_range": {"start": reps, "end": None},
                }
            )

        exercises_payload.append(
            {
                "exercise_template_id": hevy_id,
                "superset_id": None,
                "rest_seconds": None,  # rest stays in notes per your format
                "notes": ex["notes"],
                "sets": sets,
            }
        )

    return {
        "routine": {
            "title": routine["routine_title"],
            "folder_id": None,
            "notes": f"Jeff Nippard Min-Max 4x · Week {routine['week']} · {routine['day_name']}",
            "exercises": exercises_payload,
        }
    }


# --------------------------------------------------------------------
# API CALL
# --------------------------------------------------------------------
def create_routine_in_hevy(payload):
    url = BASE_URL + ROUTINES_ENDPOINT
    resp = requests.post(url, headers=HEADERS, json=payload)
    resp.raise_for_status()
    return resp.json()


def main():
    routines = build_all_routines()
    print(f"Found {len(routines)} routines total:")
    for r in routines:
        print(f"  - {r['routine_title']} ({len(r['exercises'])} exercises)")

    for routine in routines:
        payload = routine_to_hevy_payload(routine)

        if DRY_RUN:
            print("\n--- DRY RUN ---")
            print(f"Would create routine: {routine['routine_title']}")
            # Uncomment to inspect JSON:
            # import json
            # print(json.dumps(payload, indent=2))
        else:
            print(f"\nCreating routine: {routine['routine_title']}")
            result = create_routine_in_hevy(payload)
            print("Created routine, API response:", result)


if __name__ == "__main__":
    main()
